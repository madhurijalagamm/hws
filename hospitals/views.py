from django.shortcuts import render
import plotly.graph_objs as go
from plotly.offline import plot
import json
import openpyxl
from collections import Counter

def barchart_canteen():
    workbook = openpyxl.load_workbook('hospitals/survey.xlsx')
    worksheet = workbook['Form Responses 1']
    canteen_values = []
    for row in worksheet.iter_rows(min_row=2, min_col=6, max_col=6):
        for cell in row:
            canteen_values.append(cell.value)
            
            
    canteen_values = [x for x in canteen_values if x is not None]
    canteen_counts = Counter(canteen_values)
    canteen_counts = dict(sorted(canteen_counts.items()))

    # Example data
    numbers = canteen_values

    # Create a Plotly chart
    data = [go.Bar(x= ['0','1','2','3','4','5','6','7'], y=list(canteen_counts.values()))]
    layout = go.Layout(title='Number of times students go to canteen')
    fig = go.Figure(data=data, layout=layout)
    fig.update_xaxes(title_text="no of times people go to canteen")
    fig.update_yaxes(title_text="no of people")
    chart = plot(fig, output_type='div')  
    return chart  


    
def chart_view(request):
    # Open the Excel workbook
    workbook = openpyxl.load_workbook('hospitals/survey.xlsx')

    # Select the worksheet where the column is located
    worksheet = workbook['Form Responses 1']

    # Create an empty list to hold the column values
    column_values = []
    

    # Iterate over the rows in the column and append each value to the list
    for row in worksheet.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            column_values.append(cell.value)

    # Print the list of column values
    print(column_values)
    column_values = [x for x in column_values if x is not None]
    value_counts = Counter(column_values)
    value_counts = dict(sorted(value_counts.items()))

    # Example data
    numbers = column_values

    # Create a Plotly chart
    data = [go.Bar(x= ['0','1','2','3','4','5','6','7'], y=list(value_counts.values()))]
    layout = go.Layout(title='Number of times students eat breakfast')
    fig = go.Figure(data=data, layout=layout)
    fig.update_xaxes(title_text="no of times people eat breakfast")
    fig.update_yaxes(title_text="no of people")
    chart = plot(fig, output_type='div')
    
    
    
    
    #veg/non-veg
    food = []
    

    # Iterate over the rows in the column and append each value to the list
    for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            food.append(cell.value)

    # Print the list of column values
    print(food)
    food = [x for x in food if x is not None]
    food_counts = Counter(food)
    food_counts = dict(sorted(food_counts.items()))

    # Example data
    numbers = food

    # Create a Plotly chart
    data = [go.Pie(labels=list(food_counts.keys()), values=list(food_counts.values()))]
    layout = go.Layout(title='food preferences')
    fig = go.Figure(data=data, layout=layout)
    fig.update_xaxes(title_text="food preferences")
    fig.update_yaxes(title_text="Number of Students")

    # Render the chart as a Plotly div string
    chart2 = plot(fig, output_type='div')
    
    
    canteen_chart = barchart_canteen()

    # Pass the chart to the template context
    return render(request, 'health_viz.html', {'barchart1': chart, 'barchart2': canteen_chart,'barchart3': chart,'piechart1': chart2, 'piechart2': chart2})





